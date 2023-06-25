import os
from openpyxl import load_workbook
from .conftest import path_res


def test_xlsx():
    workbook = load_workbook(os.path.join(path_res, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    value = sheet.cell(row=3, column=2).value
    print(value)

    assert 'Mara' in value
